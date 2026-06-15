"""Витягування доменів з URL та парсинг даних ТЗ."""

import csv
import io
import re
from urllib.parse import urlparse
from openpyxl import load_workbook
from services.sheets import find_report_sheet


def extract_domain(value):
    """Витягнути основний домен з будь-якого формату URL.

    Приклади:
        https://example.com/page/123 -> example.com
        http://www.example.com -> example.com
        example.com -> example.com
        example.com/some-page -> example.com
    Повертає None якщо значення пусте або невалідне.
    """
    if not value or not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    # Додаємо схему якщо відсутня, щоб urlparse працював
    if not re.match(r'https?://', value, re.I):
        value = 'https://' + value

    try:
        parsed = urlparse(value)
        domain = parsed.hostname
        if not domain:
            return None
        # Видаляємо www.
        if domain.startswith('www.'):
            domain = domain[4:]
        domain = domain.lower().rstrip('.')
        # Перевірка що це схоже на домен (хоча б одна точка)
        if '.' not in domain:
            return None
        return domain
    except Exception:
        return None


def _parse_rows(rows_iter, skip_header=True):
    """Спільна логіка парсингу рядків.

    Формат аркуша 'Звіт замовнику':
    - B (індекс 1) = донор (домен або URL)
    - D (індекс 3) = акцептор (URL сайту замовника)
    - I (індекс 8) = запасне посилання на статтю (повний URL — витягуємо домен як донор)

    Повертає список (acceptor_domain, donor_domain, source_column, row_number).
    """
    results = []

    for row_idx, row in enumerate(rows_iter, start=1):
        # Пропускаємо заголовок
        if row_idx == 1 and skip_header:
            continue

        # Безпечне отримання значень
        def cell(idx):
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ''

        donor_b = cell(1)    # Колонка B — донор
        acceptor_d = cell(3)  # Колонка D — акцептор URL
        donor_i = cell(8)     # Колонка I — запасне посилання

        acceptor_domain = extract_domain(acceptor_d)
        if not acceptor_domain:
            continue

        # Донор з колонки B
        donor_domain_b = extract_domain(donor_b)
        if donor_domain_b:
            results.append((acceptor_domain, donor_domain_b, 'B', row_idx))

        # Донор з колонки I (витягуємо домен з повного URL)
        donor_domain_i = extract_domain(donor_i)
        if donor_domain_i:
            results.append((acceptor_domain, donor_domain_i, 'I', row_idx))

    return results


def parse_report_xlsx(filepath):
    """Парсити XLSX файл — шукає аркуш 'Звіт замовнику'.

    Повертає (records, full_records, error_or_none).
    records: список (acceptor_domain, donor_domain, source_column, row_number)
    full_records: список (acceptor, month, donor, anchor, final_url, backup_url, row_number)
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)

    sheet_name = find_report_sheet(wb)
    if not sheet_name:
        wb.close()
        return [], [], "Аркуш 'Звіт замовнику' не знайдено — проєкт не опублікований, пропускаємо."

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = _parse_rows(rows, skip_header=True)
    full_records = _parse_full_rows(rows, skip_header=True)
    return records, full_records, None


def _parse_full_rows(rows_iter, skip_header=True):
    """Витягнути повні рядки з Звіту замовнику.

    Формат: A=Місяць, B=Донор, D=URL(акцептор), E=Анкор, F=Final URL, I=Запаска
    Пропускаємо: G=Статус, H=Індексатор, J=Індексатор

    Повертає список (acceptor_domain, month, donor, anchor, final_url, backup_url, row_number).
    """
    results = []

    for row_idx, row in enumerate(rows_iter, start=1):
        if row_idx == 1 and skip_header:
            continue

        def cell(idx):
            if idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ''

        month = cell(0)       # A — Місяць
        donor_raw = cell(1)   # B — Донор
        acceptor_url = cell(3) # D — URL (акцептор)
        anchor = cell(4)       # E — Анкор
        final_url = cell(5)    # F — Готове посилання (Final URL)
        backup_url = cell(8)   # I — Запаска

        acceptor_domain = extract_domain(acceptor_url)
        if not acceptor_domain:
            continue

        donor = donor_raw  # Залишаємо як є (домен без http)

        if not donor:
            continue

        results.append((
            acceptor_domain, month, donor, acceptor_url,
            anchor, final_url, backup_url, row_idx
        ))

    return results


def parse_excel_file(filepath):
    """Парсити завантажений Excel файл (перший аркуш або 'Звіт замовнику')."""
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Спочатку шукаємо 'Звіт замовнику'
    sheet_name = find_report_sheet(wb)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        # Якщо немає — беремо перший аркуш
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    return _parse_rows(rows, skip_header=True)


def parse_csv_content(content, filename=""):
    """Парсити CSV контент. Повертає список записів."""
    rows = list(csv.reader(io.StringIO(content)))
    return _parse_rows(rows, skip_header=True)


def parse_master_sheet_csv(content):
    """Парсити майстер-таблицю — витягнути URLs з колонки A."""
    urls = []
    reader = csv.reader(io.StringIO(content))
    for row in reader:
        if row and row[0]:
            url = row[0].strip()
            if 'docs.google.com/spreadsheets' in url:
                urls.append(url)
    return urls


def normalize_brand(value):
    """Нормалізувати назву бренду для групування.

    'Pin Up' -> 'pinup'
    'PinUP' -> 'pinup'
    'pin-up online' -> 'pinup'
    'Пін Ап' -> 'пінап'
    """
    if not value or not isinstance(value, str):
        return None
    s = value.strip()
    if not s:
        return None
    # Lowercase
    s = s.lower()
    # Видаляємо поширені суфікси/модифікатори
    remove_words = [
        'online', 'casino', 'bet', 'game', 'games', 'play', 'slots',
        'official', 'site', 'website', 'app', 'mobile', 'live',
        'офіційний', 'сайт', 'казино', 'гра', 'ігри', 'слоти',
        'официальный', 'игры',
    ]
    for word in remove_words:
        s = re.sub(r'\b' + word + r'\b', '', s)
    # Видаляємо все крім букв і цифр
    s = re.sub(r'[^a-zA-Zа-яА-ЯіІїЇєЄґҐёЁ0-9]', '', s)
    return s if s else None


def extract_titlekw_from_xlsx(filepath):
    """Витягнути значення TitleKW з аркуша ТЗ (колонка J, індекс 9).

    Повертає список (original_value, normalized).
    """
    results = []
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Шукаємо аркуш ТЗ
    ws = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'тз':
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active  # fallback на перший аркуш

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # пропускаємо заголовок
            continue
        if len(row) > 9 and row[9]:
            val = str(row[9]).strip()
            if val:
                norm = normalize_brand(val)
                if norm:
                    results.append((val, norm))

    wb.close()
    return results


def parse_master_sheet_xlsx(filepath):
    """Парсити майстер-таблицю XLSX — витягнути URLs з колонки A."""
    urls = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            url = str(row[0]).strip()
            if 'docs.google.com/spreadsheets' in url:
                urls.append(url)
    wb.close()
    return urls
